"""Define test helper functions."""

from __future__ import annotations

import contextlib
import logging
import os
from pathlib import Path
from typing import TYPE_CHECKING

import pytest
from openpyxl import load_workbook

if TYPE_CHECKING:
    from collections.abc import Generator, Sequence

    from click import BaseCommand as CliBaseCommand
    from click.testing import CliRunner
    from click.testing import Result as CliResult
    from openpyxl import Workbook
    from openpyxl.cell import Cell, MergedCell
    from openpyxl.worksheet.worksheet import Worksheet


# Set to True to have tests save output files to an isolated temporary directory
# unique to the test method. This will preserve the tests directory.
# Set to False to save files to the test source directory which may be useful for debugging.
# pytest eventually cleans up the `tmp_path`.
# THIS APPLIES TO ALL TESTS!
USE_TMP_PATH = True


# IMPORTANT!
#
# Pytest and CliRunner don't get along when it comes to the `logging` module and this
# prevents `CliRunner.invoke` from capturing messages emitted from from the `logging` methods.
#
# https://github.com/pallets/click/issues?q=is%3Aissue%20%20%20logging%20capture%20
#
# The workaround is to use the `caplog` fixture from pytest which captures log output.
# https://docs.pytest.org/en/stable/logging.html


######################################################################
# Helper functions


def cli_invoke(  # noqa: PLR0913
    cli: CliBaseCommand,
    cli_args: Sequence[str | Path],
    runner: CliRunner,
    caplog: pytest.LogCaptureFixture,
    expected_exit_code: int = 0,
    assert_warnings: bool = True,
) -> CliResult:
    """Invoke the CLI with the given arguments.

    Asserts that it exits with the expected code and without any ERROR logs emitted.

    Args:
        runner (CliRunner): The CLI runner instance.
        caplog (pytest.LogCaptureFixture): Provides access and control of log capturing.
        cli (CliBaseCommand): The CLI command.
        cli_args (list[str]): Command line arguments.
        expected_exit_code (int): The expected exit code; default is 0.
        assert_warnings (bool): If True, assert that no WARNING logs were emitted.

    Returns:
        CliResult: The result of the CLI invocation.
    """
    if "-vv" not in cli_args:
        # Set verbose so we can assert that no ERROR or WARNING were emitted.
        cli_args = ("-vvv", *cli_args)

    if "--data-dir" not in cli_args:
        # Explicitly set data directory to the current working directory
        # so output files are written to the test working directory.
        cli_args = ("--data-dir", str(Path.cwd()), *cli_args)

    if "--recommend-reports" not in cli_args:
        # Don't recommend reports (by default) so we don't get warnings
        # related to "missing" reports.
        cli_args = ("--no-recommend-reports", *cli_args)

    # Convert all arguments to strings (invoke won't accept Path objects)
    cli_args = [str(arg) for arg in cli_args]

    result = runner.invoke(cli, cli_args)

    if result.exception is not None and not isinstance(result.exception, SystemExit):
        pytest.fail(f"Exception raised: {result.exc_info}", pytrace=True)

    if result.exit_code != expected_exit_code:
        output = [f"Unexpected exit code; {result.exit_code} != {expected_exit_code}"]

        if result.exit_code == 2:  # noqa: PLR2004
            output.append(f"Invoked with args: {cli_args}")

        if result.output:
            output.append("Output:")
            output.extend(f"> {line}" for line in result.output.splitlines())

        if result.stderr:
            output.append("Stderr:")
            output.extend(f"> {line}" for line in result.stderr.splitlines())

        # pytest shows log messages by default...
        # > if caplog.messages:
        # >    output.append("Log messages:")
        # >    output.extend(caplog.messages)

        pytest.fail("\n".join(output) if output else "No output or log messages", pytrace=False)

    assert_no_caplog_errors(caplog)

    if assert_warnings:
        assert_no_caplog_warnings(caplog)

    return result


def caplog_errors(caplog: pytest.LogCaptureFixture) -> list[logging.LogRecord]:
    """Return a list of ERROR (or higher) records emitted by `logging`.

    Args:
        caplog (pytest.LogCaptureFixture): Provides access and control of log capturing.

    Returns:
        list[logging.LogRecord]: A list of warnings captured by caplog.
    """
    return [record for record in caplog.get_records("call") if record.levelno > logging.WARNING]


def caplog_warnings(caplog: pytest.LogCaptureFixture) -> list[logging.LogRecord]:
    """Return a list of WARNING records emitted by `logging`.

    Args:
        caplog (pytest.LogCaptureFixture): Provides access and control of log capturing.

    Returns:
        list[logging.LogRecord]: A list of warnings captured by caplog.
    """
    return [record for record in caplog.get_records("call") if record.levelno == logging.WARNING]


def caplog_warning_messages(caplog: pytest.LogCaptureFixture) -> Sequence[str]:
    """Return a list of WARNING messages emitted by `logging`.

    Args:
        caplog (pytest.LogCaptureFixture): Provides access and control of log capturing.

    Returns:
        Sequence[str]: A list of warnings captured by caplog.
    """
    return tuple([record.message for record in caplog_warnings(caplog)])


def assert_no_caplog_errors(caplog: pytest.LogCaptureFixture) -> None:
    """Assert that no ERROR (or higher) records were emitted by `logging`.

    Args:
        caplog (pytest.LogCaptureFixture): Provides access and control of log capturing.
    """
    errors = caplog_errors(caplog)
    if errors:
        pytest.fail(f"Logging captured {len(errors)} ERROR messages", pytrace=False)


def assert_no_caplog_warnings(caplog: pytest.LogCaptureFixture) -> None:
    """Assert that no WARNING records were emitted by `logging`.

    Args:
        caplog (pytest.LogCaptureFixture): Provides access and control of log capturing.
    """
    warnings = caplog_warnings(caplog)
    if warnings:
        pytest.fail(f"Logging captured {len(warnings)} WARNING messages", pytrace=False)


@contextlib.contextmanager
def tmp_path_cwd(
    tmp_path: Path | None,
    tmp_path_locals: Sequence[str | Path] | None = None,
) -> Generator[Path, None, None]:
    """Change the current working directory to the given path.

    When yielded, `tmp_path_locals` will contain symlink paths to the local files.

    Args:
        tmp_path (Path | None): The path to change to.
        tmp_path_locals (Sequence[str | Path] | None): [out] Files to make available in the test working directory.
            Ignored if `tmp_path` is None or `USE_TMP_PATH` is False.

    Yields:
        Path: The new current working directory.
    """

    def _local_file_abspath(local_file: Path, test_cwd: Path) -> Path:
        if not local_file.is_absolute():
            local_file = test_cwd / local_file
        return local_file.resolve() if ".." in local_file.parts else local_file

    test_cwd = runner_cwd = Path.cwd()  # Current test working directory
    use_tmp_path = tmp_path and USE_TMP_PATH and (tmp_path != test_cwd)

    if use_tmp_path:
        assert tmp_path is not None

        os.chdir(tmp_path)
        runner_cwd = tmp_path

        if tmp_path_locals is not None:
            assert isinstance(tmp_path_locals, list)

            for i, local_file in enumerate(tmp_path_locals):
                if isinstance(local_file, str):
                    local_file = Path(local_file)  # noqa: PLW2901

                src_path = _local_file_abspath(local_file, test_cwd)
                dst_path = runner_cwd / src_path.name

                assert src_path.exists(), f"Source path does not exist: {src_path}"
                assert src_path.is_file(), f"Source path is not a file: {src_path}"

                # Create symlink to the file in the test working directory.
                # If `dst_path` is a symlink, `exists` checks the target exists unless `follow_symlinks=False`.
                if not dst_path.exists() and not dst_path.exists(follow_symlinks=False):
                    os.symlink(src_path, dst_path)
                    tmp_path_locals[i] = dst_path

    try:
        yield runner_cwd
    finally:
        if use_tmp_path:
            os.chdir(test_cwd)


def assert_compare_xlsx(result_path: Path, expected_path: Path) -> None:  # noqa: C901, PLR0915
    """Compare two xlsx files.

    Args:
        result_path (Path): Path to test result workbook.
        expected_path (Path): Path to expected test result workbook.
    """

    def _compare_cell(sheetname: str, result_cell: Cell | MergedCell, expected_cell: Cell | MergedCell) -> None:
        if result_cell.value != expected_cell.value:
            pytest.fail(
                f"{sheetname}[{result_cell.coordinate}] value does not match; Expected '{expected_cell.value}' != '{result_cell.value}'",
                pytrace=False,
            )

        if repr(result_cell.fill) != repr(expected_cell.fill):
            pytest.fail(
                f"{sheetname}[{result_cell.coordinate}] fill properties do not match",
                pytrace=False,
            )

        if repr(result_cell.font) != repr(expected_cell.font):
            pytest.fail(
                f"{sheetname}[{result_cell.coordinate}] font properties do not match",
                pytrace=False,
            )

        if result_cell.number_format != expected_cell.number_format:
            pytest.fail(
                f"{sheetname}[{result_cell.coordinate}] number_format does not match; "
                f"Expected '{expected_cell.number_format}' != '{result_cell.number_format}'",
                pytrace=False,
            )

        if result_cell.style != expected_cell.style:
            pytest.fail(
                f"{sheetname}[{result_cell.coordinate}] style does not match; Expected '{expected_cell.style}' != '{result_cell.style}'",
                pytrace=False,
            )

    def _compare_columns(sheetname: str, result_ws: Worksheet, expected_ws: Worksheet) -> None:
        result_columns = [cells[0].value for cells in result_ws.iter_cols(max_row=1)]
        expected_columns = [cells[0].value for cells in expected_ws.iter_cols(max_row=1)]

        if result_columns != expected_columns:
            result_columns = set(result_columns)
            expected_columns = set(expected_columns)

            if expected_columns - result_columns:
                pytest.fail(f"'{sheetname}' missing columns: {expected_columns - result_columns}", pytrace=False)
            elif result_columns - expected_columns:
                pytest.fail(f"'{sheetname}' unexpected columns: {result_columns - expected_columns}", pytrace=False)
            else:
                pytest.fail(f"'{sheetname}' column order does not match; Expected {expected_columns} != {result_columns}", pytrace=False)

        assert result_ws.max_column == expected_ws.max_column

    def _compare_rows(sheetname: str, result_ws: Worksheet, expected_ws: Worksheet) -> None:
        if result_ws.max_row != expected_ws.max_row:
            pytest.fail(f"'{sheetname}' sheet row count does not match: Expected {expected_ws.max_row} != {result_ws.max_row}", pytrace=False)

        for row in range(1, result_ws.max_row + 1):
            for col in range(1, result_ws.max_column + 1):
                _compare_cell(sheetname, result_ws.cell(row=row, column=col), expected_ws.cell(row=row, column=col))

    def _compare_sheet_names(result_wb: Workbook, expected_wb: Workbook) -> None:
        if result_wb.sheetnames != expected_wb.sheetnames:
            result_sheets = set(result_wb.sheetnames)
            expected_sheets = set(expected_wb.sheetnames)

            if expected_sheets - result_sheets:
                pytest.fail(f"Missing sheets: {expected_sheets - result_sheets}", pytrace=False)
            elif result_sheets - expected_sheets:
                pytest.fail(f"Unexpected sheets: {result_sheets - expected_sheets}", pytrace=False)
            else:
                pytest.fail(f"Sheet order does not match; Expected {expected_wb.sheetnames} != {result_wb.sheetnames}", pytrace=False)

    try:
        result_wb = load_workbook(result_path)
    except Exception as exc:  # noqa: BLE001
        pytest.fail(f"Failed to load workbook: {result_path}; {exc}", pytrace=False)

    try:
        expected_wb = load_workbook(expected_path)
    except Exception as exc:  # noqa: BLE001
        pytest.fail(f"Failed to load workbook: {expected_path}; {exc}", pytrace=False)

    _compare_sheet_names(result_wb, expected_wb)

    for sheetname in result_wb.sheetnames:
        result_ws = result_wb[sheetname]
        expected_ws = expected_wb[sheetname]

        _compare_columns(sheetname, result_ws, expected_ws)
        _compare_rows(sheetname, result_ws, expected_ws)


def create_and_compare_workbook(  # noqa: PLR0913
    cli: CliBaseCommand,
    runner: CliRunner,
    tmp_path: Path,
    resources_dir: Path,
    caplog: pytest.LogCaptureFixture,
    tmp_path_locals: Sequence[str | Path] | None = None,
    expected_exit_code: int = 0,
    assert_warnings: bool = True,
) -> CliResult:
    """Create workbook `result.xlsx` and compare it with `expected_result.xlsx`.

    Uses the main configuration file `mnyxls_tests.yaml` and local workbook configuration file `mnyxls_workbook.yaml`.

    Args:
        cli (CliBaseCommand): The CLI command.
        runner (CliRunner): The CLI runner instance.
        tmp_path (Path): Temporary directory pytest creates for the test.
        resources_dir (Path): Path to test resources directory.
        caplog (pytest.LogCaptureFixture): Provides access and control of log capturing.
        tmp_path_locals (Sequence[str | Path] | None): [out] Files to make available in the test working directory.
            Ignored if `tmp_path` is None or `USE_TMP_PATH` is False.
        expected_exit_code (int): The expected exit code; default is 0.
        assert_warnings (bool): If True, assert that no WARNING logs were emitted.

    Returns:
        CliResult: The result of the CLI invocation.
    """
    test_cwd = Path.cwd()
    result_name = "result.xlsx"

    # These files must be accessible in `runner_cwd`
    if tmp_path_locals is None:
        tmp_path_locals = []
    tmp_path_locals = [*tmp_path_locals, "mnyxls_workbook.yaml"]

    cli_args: list[str | Path] = [
        "--config-file",
        resources_dir / "mnyxls_tests.yaml",
        "--db-file",
        resources_dir / "sample_reports/txns_2018_2022.sqlite3",
        "--xls-file",
        result_name,
        "--no-reports",
    ]

    with tmp_path_cwd(tmp_path, tmp_path_locals) as runner_cwd:
        result = cli_invoke(
            cli,
            cli_args,
            runner,
            caplog,
            expected_exit_code=expected_exit_code,
            assert_warnings=assert_warnings,
        )

        if result.exit_code == 0:
            result_path = runner_cwd / result_name
            assert result_path.exists(), "Expected an .xlsx file"

            out_files = list(runner_cwd.glob("*.sqlite3"))
            assert not out_files, "Expected no .sqlite3 files"

            assert_compare_xlsx(result_path, test_cwd / "expected_result.xlsx")

    return result

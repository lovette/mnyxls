from __future__ import annotations

import contextlib
from datetime import date
from types import MappingProxyType
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import PatternFill

from .configtypes import WorksheetConfigOptionsTxnsPivotT, WorksheetConfigSelectTxnsPivotT
from .dbschema import TABLE_ACCOUNTS, TABLE_CATEGORIES, TABLE_ERAS, TABLE_TXNS, table_schema_columns
from .dbsqlite import db_list_eras
from .dbviews import VIEW_TXNS_WITHTYPEANDCLASS
from .mysqlstmt_selectand import SelectAnd
from .shared import MnyXlsConfigError, MnyXlsRuntimeError, pd_read_sql, validate_config_typed_dict
from .worksheet import (
    FORMAT_ALIGN_LEFT,
    FORMAT_BOLD_FONT,
    FORMAT_NORMAL_FONT,
    WORKSHEET_COLWIDTH_MAX,
    DataFrameKeyT,
    MoneyWorksheet,
)
from .worksheet_txns_base import MoneyWorksheetTxnsBase

if TYPE_CHECKING:
    import sqlite3
    from collections.abc import Sequence

    from .configtypes import WorksheetConfigSelectT
    from .workbook import MoneyWorkbook


# Alternating fill colors for multi-level headers.
# > https://openpyxl.readthedocs.io/en/stable/styles.html#fill
FORMAT_FILLS = [
    (
        PatternFill(start_color=c1, fill_type="solid"),
        PatternFill(start_color=c2, fill_type="solid"),
    )
    for c1, c2 in (
        # Light tones
        # > ("F0F0F0", "F8F8F8"),
        # > ("E0F2FE", "E8F5E9"),
        # > ("E3F2FD", "E8F5E9"),
        # Shades of darker blue
        # > ("3A7CA5", "4682B4"),
        # > ("6495ED", "5B92E5"),
        # > ("4F7CAA", "517CB7"),
        # Shades of blue
        # > ("E6F3FF", "E6ECFF"),
        # > ("E6E6FF", "F0F4FF"),
        # > ("E6F0FF", "F2F6FF"),
        # Pastel colors
        ("E6F2FF", "E6F3E6"),
        ("F0E6FF", "FFF0E6"),
        ("E6FFF0", "FFE6E6"),
    )
]

# Valid date-related options for `options.rows` and `options.columns`
# configuration directive *ordered by hierarchy*.
# > Mapping[directive] = "ColumnName"
PIVOT_DATEOPT_COL_MAP = MappingProxyType(
    {
        "yyyy": "DateYYYY",  # Group by year
        "yyyymm": "DateYYYYMM",  # Group by year and month
    }
)

# Valid options for `options.rows` configuration directive *ordered by hierarchy*.
# > Mapping[directive] = "ColumnName"
PIVOT_INDEX_OPT_COL_MAP = MappingProxyType(
    {
        "era": "EraName",
        "account_classification": "AccountClassification",
        "account_category": "AccountCategory",
        "account": "Account",
        "txntype": "TxnType",
        "txnclass": "TxnClass",
        "category": "Category",
        "category_subcategory": "CategorySubcategory",
        "subcategory": "Subcategory",
        "payee": "Payee",
    }
    | PIVOT_DATEOPT_COL_MAP
)


# Valid options for `options.columns` configuration directive *ordered by hierarchy*.
# > Mapping[directive] = "ColumnName"
PIVOT_COLUMNS_OPT_COL_MAP = MappingProxyType(
    {
        "era": "EraName",
        "account_classification": "AccountClassification",
        "account_category": "AccountCategory",
        "txntype": "TxnType",
        "txnclass": "TxnClass",
        "category": "Category",
    }
    | PIVOT_DATEOPT_COL_MAP
)

PIVOT_INDEX_OPT_COL_MAP_REVERSED = MappingProxyType({v: k for k, v in PIVOT_INDEX_OPT_COL_MAP.items()})
PIVOT_COLUMNS_OPT_COL_MAP_REVERSED = MappingProxyType({v: k for k, v in PIVOT_COLUMNS_OPT_COL_MAP.items()})
PIVOT_COL_MAP_OPT = PIVOT_INDEX_OPT_COL_MAP_REVERSED | PIVOT_COLUMNS_OPT_COL_MAP_REVERSED

# > Mapping[directive] = (strftime, excel_format)
PIVOT_DATEOPT_FORMATS = MappingProxyType(
    {
        "DateYYYY": (
            "strftime('%Y-12-31', `Date`)",  # strftime
            ("yyyy", None),  # Excel format (top-level, nested)
        ),
        "DateYYYYMM": (
            "strftime('%Y-%m-01', `Date`)",
            ("yyyy-mm", "mmm"),
        ),
    }
)

PIVOT_OPT_YYYY, PIVOT_COLUMN_YYYY = next(iter(PIVOT_DATEOPT_COL_MAP.items()))

# Ensure all date options have formats
assert set(PIVOT_DATEOPT_COL_MAP.values()) == set(PIVOT_DATEOPT_FORMATS.keys())

PIVOT_OPT_TOTAL = "total"

######################################################################
# Categories worksheet


class MoneyWorksheetTxnsPivot(MoneyWorksheetTxnsBase):
    """Worksheet for transactions pivot table."""

    SHEET_TYPE = "txns:pivot"

    # Keep column widths within reasonable bounds
    WORKSHEET_COLWIDTH_BOUNDS = MappingProxyType(
        {
            "CategorySubcategory": (20, WORKSHEET_COLWIDTH_MAX),
        }
    )

    def __init__(self, *args, **kwargs) -> None:  # noqa: D107
        super().__init__(*args, **kwargs)

        self._date_col: str = "Date"
        self._select_columns: list[str] = []
        self._pivot_index_names = [PIVOT_INDEX_OPT_COL_MAP[opt] for opt in ("category", "subcategory")]
        self._pivot_columns_names = [PIVOT_COLUMNS_OPT_COL_MAP[PIVOT_OPT_YYYY]]

    ######################################################################
    # Class methods

    @classmethod
    def validate_config_select(cls, workbook: MoneyWorkbook, select_config: WorksheetConfigSelectT, config_keys: str | Sequence[str]) -> None:
        """Validate worksheet configuration `select` directive.

        Args:
            workbook (MoneyWorkbook): Workbook.
            select_config (WorksheetConfigSelectT): Select configuration.
            config_keys (str | Sequence[str]): Configuration keys for reporting errors.

        Raises:
            MnyXlsConfigError: Invalid configuration.
        """
        MoneyWorksheet.validate_config_select(workbook, select_config, config_keys)

        validate_config_typed_dict(select_config, WorksheetConfigSelectTxnsPivotT, workbook.config, config_keys)

    ######################################################################
    # Instance methods

    def prepare_write_sheet(self, conn: sqlite3.Connection, writer: pd.ExcelWriter) -> None:  # noqa: C901
        """Prepare to write worksheet to workbook.

        Invoked prior to `write_sheet`.

        Args:
            conn (sqlite3.Connection): SQLite connection.
            writer (pd.ExcelWriter): Excel writer.

        Raises:
            MnyXlsConfigError
        """

        def _validate_options(options: Sequence[str], valid_options: Sequence[str], config_key: str) -> None:
            """Validate options are valid."""
            invalid_options = set(options) - set(valid_options)
            if invalid_options:
                invalid_options = "', '".join(sorted(invalid_options))
                raise MnyXlsConfigError(
                    f"Invalid options: '{invalid_options}'",
                    self.workbook.config,
                    (*self.config_keys, "options", config_key),
                )

        def _remove_duplicates(options: Sequence[str]) -> list[str]:
            """Remove duplicates while preserving order."""
            seen = set()
            unique_options = []

            for option in options:
                if option not in seen:
                    seen.add(option)
                    unique_options.append(option)

            return unique_options

        def _adjust_options(options: list[str]) -> None:
            """Adjust options to ensure valid combinations."""
            if "categorysubcategory" in options:
                # Remove individual Category and Subcategory options
                with contextlib.suppress(ValueError):
                    options.remove("category")
                    options.remove("subcategory")
            elif "subcategory" in options and "category" not in options:
                # Subcategory requires Category
                options.append("category")

        def _ensure_option_order(options: Sequence[str], valid_options: Sequence[str]) -> tuple[str, ...]:
            """Ensure options are in the correct order."""
            return tuple(sorted(options, key=lambda x: valid_options.index(x)))

        super().prepare_write_sheet(conn, writer)

        pivot_options: WorksheetConfigOptionsTxnsPivotT = self.worksheet_config.get("options", {})

        # Validate and adjust row and column "options"
        for config_key, valid_options in (
            ("rows", tuple(PIVOT_INDEX_OPT_COL_MAP.keys())),
            ("columns", (*tuple(PIVOT_COLUMNS_OPT_COL_MAP.keys()), PIVOT_OPT_TOTAL)),
        ):
            config_values = pivot_options.get(config_key)
            if not config_values:
                continue

            config_values = [config_values] if isinstance(config_values, str) else _remove_duplicates(config_values)

            _validate_options(config_values, valid_options, config_key)
            _adjust_options(config_values)

            pivot_options[config_key] = _ensure_option_order(config_values, valid_options)

        if "rows" in pivot_options:
            self._pivot_index_names = [PIVOT_INDEX_OPT_COL_MAP[opt] for opt in pivot_options["rows"]]

        if "columns" in pivot_options:
            if PIVOT_OPT_TOTAL not in pivot_options["columns"]:
                self._pivot_columns_names = [PIVOT_COLUMNS_OPT_COL_MAP[opt] for opt in pivot_options["columns"]]
            elif set(self._pivot_index_names) & set(self._pivot_columns_names):
                # We need to change the pivot column to be distinct from chosen index columns.
                for opt, col in PIVOT_COLUMNS_OPT_COL_MAP.items():
                    if col not in self._pivot_index_names and opt not in PIVOT_DATEOPT_COL_MAP:
                        self._pivot_columns_names = [col]
                        break

        # Ensure pivot columns and pivot rows do not overlap
        overlap_columns = set(self._pivot_index_names) & set(self._pivot_columns_names)
        if overlap_columns:
            directives = sorted([PIVOT_COL_MAP_OPT.get(col, col) for col in overlap_columns])

            raise MnyXlsConfigError(
                f"Pivot columns and pivot rows cannot overlap; both contain {directives}",
                self.workbook.config,
                (*self.config_keys, "options"),
            )

        self._select_columns = [*self._pivot_columns_names, *self._pivot_index_names]

    def get_sheet_data(self, conn: sqlite3.Connection) -> pd.DataFrame:  # noqa: C901, PLR0912
        """Query database and return data to write to worksheet.

        ```
        "Expenses by Category":
            sheet_type: "txns:pivot"
            select:
                txntype: "Expense"
        ```

        Args:
            conn (sqlite3.Connection): SQLite connection.
        """
        assert self.workbook is not None
        assert self.workbook.config is not None

        date_cols: list[str] = []
        select_columns = self._select_columns.copy()
        groupby_columns = self._select_columns.copy()
        txns_view_columns = (*table_schema_columns(TABLE_TXNS), *table_schema_columns(TABLE_CATEGORIES))
        pivot_options: WorksheetConfigOptionsTxnsPivotT = self.worksheet_config.get("options", {})
        opt_columns_pivot_total = PIVOT_OPT_TOTAL in pivot_options.get("columns", {})
        era_names = db_list_eras(conn) if "EraName" in select_columns else None

        q_select = SelectAnd(VIEW_TXNS_WITHTYPEANDCLASS, named="T")

        if "AccountCategory" in select_columns or "AccountClassification" in select_columns:
            q_select.join(TABLE_ACCOUNTS, ".Account")

        if era_names:
            # Join the Eras table to get Eras.rowid for sorting
            q_select.left_join(TABLE_ERAS, ".EraName")

        # Select date columns
        for date_col in PIVOT_DATEOPT_COL_MAP.values():
            if date_col in select_columns:
                date_cols.append(date_col)
                q_select.column_expr(PIVOT_DATEOPT_FORMATS[date_col][0], named=date_col)
                select_columns.remove(date_col)  # field is now a column_expr

        # Qualify column names (so `Account` isn't ambiguous)
        for i, col in enumerate(select_columns):
            if col in txns_view_columns:
                select_columns[i] = f"T.{select_columns[i]}"
        for i, col in enumerate(groupby_columns):
            if col in txns_view_columns:
                groupby_columns[i] = f"T.{groupby_columns[i]}"

        q_select.column(select_columns)
        q_select.column_expr("SUM(`Amount`)", named="Amount")

        self.apply_txns_select_where(conn, q_select, self.worksheet_config)

        q_select.group_by(groupby_columns)

        orderby_columns = groupby_columns.copy()  # used for final ORDER BY

        if era_names:
            # Order Eras by date range (though this has no effect on the actual pivot table itself)
            orderby_columns = ["Eras.rowid" if "EraName" in col else col for col in orderby_columns]

        q_select.order_by(orderby_columns)

        df_worksheet = pd_read_sql(
            conn,
            q_select,
            date_cols=dict.fromkeys(date_cols, "%Y-%m-%d") if date_cols else None,
            currency_cols=["Amount"],
        )

        if df_worksheet.empty:
            return df_worksheet

        # Pivot table drops NA values, which we have in cases such as a Category without Subcategories.
        # `pivot_table` has the `dropna=False` argument, but it still seems like rows get dropped
        # when we don't want them to.
        for col in df_worksheet.columns:
            if col not in ("Amount", *PIVOT_DATEOPT_COL_MAP.values()):
                df_worksheet[col] = df_worksheet[col].fillna("")

        if era_names:
            # Pivot tables sorted alphabetically and we want Eras to be
            # sorted by date range (which is the order in the database.)
            # The easiest way is to change the column type to an ordered Categorical.
            assert "EraName" in df_worksheet.columns
            df_worksheet["EraName"] = pd.Categorical(df_worksheet["EraName"], categories=era_names, ordered=True)

        # The `DateGroup` column is a datetime.date object which means the column headers of the pivot
        # table will be as well and we need to keep it that way so `to_excel` can write headers as a date.
        # If we convert it to a string, it will be written as text and Excel will not be able to format it as a date.
        # NO > df_worksheet[date_col] = df_worksheet[date_col].apply(lambda x: x.strftime("%Y-%m-%d"))

        # Group into columns
        try:
            df_worksheet = df_worksheet.pivot_table(
                index=self._pivot_index_names,
                columns=self._pivot_columns_names,
                values="Amount",
                aggfunc=lambda x: x.sum(min_count=1),  # Using "sum" results in "0" instead of "NaN" when Categorical columns are present...
                margins=True,
                margins_name="Total",
                observed=True,  # Only show Categorical values that are in the data.
            )
        except ValueError as err:
            raise MnyXlsRuntimeError(f"Pivot failed: {err}") from err

        # If using `dropna=False` we need to drop rows with no Total
        # > df_worksheet = df_worksheet[df_worksheet["Total"].notna()]

        # Reset index to make "index column" an actual column again
        # > df_worksheet = df_worksheet.reset_index()

        if opt_columns_pivot_total:
            # Return only the last column (which is the "Total" column)
            return df_worksheet.iloc[:, -1:]

        return df_worksheet

    def prepare_to_excel(self) -> None:
        """Prepare to write worksheet to workbook.

        Invoked prior to `to_excel`.
        `self.df_worksheet` is set to the result of `get_sheet_data`.

        Returns:
            pd.DataFrame: DataFrame to write to workbook.
        """
        assert self.df_worksheet is not None

        # Freeze (# rows, # columns)
        self.freeze_panes = (self.header_levels, self.index_levels)

        # Let `super().prepare_to_excel` know we want to avoid Excel marking currency cells
        # with "Number Stored as Text" (best we can.)
        self.currency_cols = []

        if self.header_levels == 1:
            assert isinstance(self.df_worksheet.columns, pd.Index)
            for col in self.df_worksheet.columns:
                if col not in self._pivot_columns_names:
                    self.currency_cols.append(col)
        else:
            # All columns are currency because the non-currency columns are in the index
            # so are not enumerated in `.columns`.
            assert isinstance(self.df_worksheet.columns, pd.Index | pd.MultiIndex)
            for col in self.df_worksheet.columns:
                self.currency_cols.append(col)

        super().prepare_to_excel()

    def format_sheet(self) -> None:  # noqa: C901, PLR0912
        """Format worksheet after sheet is created."""
        super().format_sheet()

        assert self.pyxl_worksheet is not None
        assert self.df_worksheet is not None

        options = self.worksheet_config.get("options", {})

        # The left-hand `index_levels` columns are the index columns.
        # We override some of the `header_style` formatting.
        for col_cell in self.iter_index_cols():
            for row_cell in self.iter_header_rows(cell=col_cell):
                row_cell.alignment = FORMAT_ALIGN_LEFT

            for row_cell in self.iter_data_rows(cell=col_cell):
                row_cell.alignment = FORMAT_ALIGN_LEFT
                row_cell.font = FORMAT_NORMAL_FONT

        # Format date index columns as dates.
        # date_level depicts hierarchy: YYYY -> YYYYMM
        date_level = 0
        for date_col, row_cell in self._get_date_index_cols().items():
            for col_cell in self.iter_data_rows(cell=row_cell):
                assert isinstance(col_cell, Cell)
                if isinstance(col_cell.value, date):  # "Total" row is not a date
                    fmt = PIVOT_DATEOPT_FORMATS[date_col][1][date_level]
                    if fmt is not None:
                        col_cell.number_format = fmt

            date_level += 1  # noqa: SIM113

        # Non-index columns are currency amounts.
        for col_cell in self.iter_data_cols():
            for row_cell in self.iter_data_rows(cell=col_cell):
                row_cell.number_format = self.amount_number_format

        # Format date header rows as dates.
        # date_level depicts hierarchy: YYYY -> YYYYMM
        date_level = 0
        for date_col, row_cell in self._get_date_header_rows().items():
            for col_cell in self.iter_data_cols(cell=row_cell):
                assert isinstance(col_cell, Cell)
                if isinstance(col_cell.value, date):  # "Total" column is not a date
                    fmt = PIVOT_DATEOPT_FORMATS[date_col][1][date_level]
                    if fmt is not None:
                        col_cell.number_format = fmt

            date_level += 1

        # Bold the "Total" row
        total_row_cell = self.pyxl_worksheet["A"][-1]
        if total_row_cell is not None and total_row_cell.value == "Total":
            for row_cells in self.pyxl_worksheet.iter_rows(min_row=total_row_cell.row, max_row=total_row_cell.row):
                for row_cell in row_cells:
                    assert isinstance(row_cell, Cell)
                    row_cell.font = FORMAT_BOLD_FONT

        # Alternate fill colors of multi-level headers.
        if options.get("colorful", True):
            for row_idx, row_cell in enumerate(self.iter_header_rows(letter="A")):
                assert row_cell.row is not None
                alternate_fills = FORMAT_FILLS[row_idx % len(FORMAT_FILLS)]
                if row_cell.row <= self.df_worksheet.columns.nlevels:  # don't highlight Index columns headers
                    for idx, col_cell in enumerate(self.iter_data_cols(cell=row_cell), 1):
                        assert isinstance(col_cell, Cell)
                        assert col_cell.column is not None
                        if col_cell.column < self.pyxl_worksheet.max_column:  # don't highlight "Total" column
                            col_cell.fill = alternate_fills[idx % 2]

    def autofit_columns(self, width_hints: dict[DataFrameKeyT, int] | None = None) -> None:
        """Autofit column widths based on maximum value string lengths.

        Fails until `to_excel` is complete.
        Requires worksheet to have a header row.
        Column headers that do not exist in the worksheet are ignored.
        Column widths are set to the maximum value string length by default.

        Args:
            width_hints (dict[str, int]): Column width hints by column header.
                Sheet is not required to honor these hints.
        """
        assert self.currency_cols is not None

        # Autofit for this sheet type is different than other sheets in
        # that the width of the largest currency is applied to all other currency columns.
        # `width_hints` is changed in place, so the first sheet of this
        # type will dictate the widths for all sheets of this type.

        if width_hints is not None:
            max_currency_len = 0

            for col_df_idx, width in width_hints.items():
                if col_df_idx in self.currency_cols:
                    max_currency_len = max(max_currency_len, width)

            for col_df_idx in width_hints:
                if col_df_idx in self.currency_cols:
                    width_hints[col_df_idx] = max_currency_len

        super().autofit_columns(width_hints)

    ######################################################################
    # Helper methods

    def _get_date_header_rows(self) -> dict[str, Cell | MergedCell]:
        """Get cells referencing header rows that are dates.

        Returns:
            Sequence[Cell]
        """
        assert self.pyxl_worksheet is not None

        date_header_rows: dict[str, Cell | MergedCell] = {}

        if self.header_levels == 1:
            # We have a single header row, so the date headers are on the first row.
            if self._pivot_columns_names[0] in PIVOT_DATEOPT_FORMATS:
                date_header_rows[self._pivot_columns_names[0]] = self.pyxl_worksheet["A1"]
        else:
            # We have to search all header index columns for the name because it will change
            # depending on how many index and column levels there are.
            for header_col in self.iter_index_cols():
                for row_cell in self.iter_header_rows(cell=header_col):
                    if isinstance(row_cell.value, str) and row_cell.value in PIVOT_DATEOPT_FORMATS:
                        date_header_rows[row_cell.value] = row_cell

        return date_header_rows

    def _get_date_index_cols(self) -> dict[str, Cell | MergedCell]:
        """Get cells referencing index columns that are dates.

        Returns:
            Sequence[Cell]
        """
        assert self.pyxl_worksheet is not None

        date_index_cols: dict[str, Cell | MergedCell] = {}

        if self.index_levels == 1:
            # We have a single index column, so the date headers are in the first column.
            if self._pivot_index_names[0] in PIVOT_DATEOPT_FORMATS:
                date_index_cols[self._pivot_index_names[0]] = self.pyxl_worksheet["A1"]
        else:
            # We have to search all header index columns for the name because it will change
            # depending on how many index and column levels there are.
            for header_col in self.iter_index_cols():
                for row_cell in self.iter_header_rows(cell=header_col):
                    if isinstance(row_cell.value, str) and row_cell.value in PIVOT_DATEOPT_FORMATS:
                        date_index_cols[row_cell.value] = row_cell

        return date_index_cols


######################################################################
# Register worksheet class so it can be referenced in the configuration file.

MoneyWorksheetTxnsPivot.register_sheet_type()

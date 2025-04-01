from __future__ import annotations

import logging
import typing
from collections import defaultdict
from datetime import date
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import load_workbook

from .configtypes import (
    ConfigFileValueT,
    MainConfigFileT,
    WorkbookConfigFileT,
    WorkbookConfigSelectT,
    WorkbookConfigT,
    WorksheetConfigSelectT,
)
from .shared import (
    MnyXlsConfigError,
    config_warning,
    get_date_relative_to,
    get_select_values,
    parse_yyyymmdd_flex,
    read_config_file,
    resolve_rel_path,
    validate_config_typed_dict,
)
from .worksheet import DataFrameKeyT, MoneyWorksheet

# Import all worksheet classes so they register themselves with `register_sheet_type`.
from .worksheet_accounts import MoneyWorksheetAccounts  # noqa: F401
from .worksheet_categories import MoneyWorksheetCategories  # noqa: F401
from .worksheet_naked_categories import MoneyWorksheetCategoriesNaked  # noqa: F401
from .worksheet_singlepayee_categories import MoneyWorksheetCategoriesSinglePayee  # noqa: F401
from .worksheet_txns import MoneyWorksheetTxns  # noqa: F401
from .worksheet_txns_pivot import MoneyWorksheetTxnsPivot  # noqa: F401

if TYPE_CHECKING:
    import sqlite3
    from collections.abc import Sequence
    from pathlib import Path

    import click

    from .configtypes import WorksheetConfigT

WORKBOOK_MAX_SHEETS = 75  # Arbitrary limit
TEMPORARY_SHEET_NAME = "No sheets were generated"[:31]
DEFAULT_XLS_TEMPLATE_NAME = "mnyxls_template.xlsx"
DEFAULT_XLS_CONFIG_NAME = "mnyxls_workbook.yaml"

DEFAULT_WORKBOOK_CONFIG: WorkbookConfigT = {
    "worksheets": {
        "Accounts": {
            "sheet_type": "accounts",
        },
        "Transactions": {
            "sheet_type": "txns",
            "foreach": "txntype",
        },
        "Category by year": {
            "sheet_type": "txns:pivot",
            "options": {
                "rows": ["category", "subcategory"],
                "columns": "yyyy",
            },
        },
    },
}

# Use a named logger instead of root logger
logger = logging.getLogger("mnyxls")


######################################################################
# MoneyWorkbook


class MoneyWorkbook:
    """Mnyxls workbook."""

    def __init__(self, xls_file: Path, xls_config: WorkbookConfigFileT | None, config: MainConfigFileT) -> None:
        """Constructor.

        Args:
            xls_file (Path): Path to workbook.
            xls_config (WorkbookConfigFileT|None): Workbook configuration.
            config (ConfigFileT): Top-level user configuration directives.
        """
        assert config is not None

        data_dir: Path | None = config.get("data_dir")
        xls_template: Path | None = config.get("xls_template")

        assert data_dir is not None

        self.xls_file = xls_file
        self.xls_template = resolve_rel_path(xls_template, data_dir, DEFAULT_XLS_TEMPLATE_NAME)
        self.config = xls_config or config  # top-level or ancillary config file?
        self.worksheet_specs: list[tuple[type[MoneyWorksheet], str, WorksheetConfigT]] = []
        self.worksheets: list[MoneyWorksheet] = []  # Use a list so requested sheet names don't have to be unique
        self.empty_worksheets: list[MoneyWorksheet] = []
        self.can_use_existing_sheet: dict[str, bool] = {}

        # Workbook configuration can be specified in toplevel config or a separate config file.
        # If neither is specified, use the default configuration.
        self.workbook_config = self.config.get("workbook", DEFAULT_WORKBOOK_CONFIG)

    def __repr__(self) -> str:
        """Return a string representation of the object for debugging purposes.

        Returns:
            str
        """
        return f"{self.__class__.__name__}('{self.xls_file.name}')"

    ######################################################################
    # Instance methods

    def get_config_value(self, key: str, default: ConfigFileValueT | None) -> ConfigFileValueT | None:
        """Get a workbook configuration value.

        Workbook configuration values override top-level configuration values.

        Args:
            key (str): Configuration key.
            default (str): Default value if key is not set.

        Returns:
            ConfigFileValueT | None
        """
        return self.workbook_config.get(key, self.config.get(key, default))

    def validate_config(self) -> None:
        """Validate workbook configuration.

        Raises:
            MnyXlsConfigError: Invalid configuration.
        """
        # Workbook configuration was set in the constructor, we're just validating it here.
        assert self.workbook_config is not None

        if "workbook" in self.config:
            logger.debug(f"Workbook config path: '{self.config.get('_config_file')}'")
            validate_config_typed_dict(self.workbook_config, WorkbookConfigT, self.config, "workbook")
        else:
            logger.debug("Using default workbook configuration.")

        config_worksheets: dict[str, WorksheetConfigT] = self.workbook_config.get("worksheets")
        workbook_select: WorkbookConfigSelectT | None = self.workbook_config.get("select")

        if workbook_select:
            self.validate_config_select(self, workbook_select, ("workbook", "select"))

        assert config_worksheets is not None

        # Worksheet "config key" is the sheet name or a placeholder for a 'foreach' directive.
        for sheet_config_key in config_worksheets:  # noqa: PLC0206
            config_worksheets[sheet_config_key]["_config_key"] = sheet_config_key

        for sheet_config_key, worksheet_config in config_worksheets.items():
            config_keys = ("workbook", "worksheets", sheet_config_key)

            worksheet_type = worksheet_config.get("sheet_type")
            if not worksheet_type:
                raise MnyXlsConfigError("Directive is required.", self.config, (*config_keys, "sheet_type"))

            worksheet_cls = MoneyWorksheet.get_worksheet_class(worksheet_type)
            if worksheet_cls is None:
                raise MnyXlsConfigError("Unrecognized sheet type.", self.config, (*config_keys, worksheet_type))

            worksheet_cls.validate_config(self, sheet_config_key, worksheet_config)

            if workbook_select:
                # Apply workbook select directives to all worksheets.
                # Worksheet directives override workbook directives.
                # Merge *after* worksheet validation so any warnings about the workbook directives
                # are not repeated for the worksheet directives.
                worksheet_select = worksheet_config.get("select", {})
                worksheet_config["select"] = typing.cast("WorksheetConfigSelectT", workbook_select) | worksheet_select

            self.worksheet_specs.append((worksheet_cls, sheet_config_key, worksheet_config))

    @classmethod
    def validate_config_select(
        cls,
        workbook: MoneyWorkbook,
        select_config: WorkbookConfigSelectT | WorksheetConfigSelectT,
        config_keys: str | Sequence[str],
    ) -> None:
        """Validate workbook or worksheet configuration `select` directive.

        Worksheets use the same `select` directive as the workbook, so this method is shared
        so validation is consistent.

        Args:
            workbook (MoneyWorkbook): Workbook.
            select_config (WorkbookConfigSelectT | WorksheetConfigSelectT): Select configuration.
            config_keys (str | Sequence[str]): Configuration keys for reporting errors.

        Raises:
            MnyXlsConfigError: Invalid configuration.
        """

        def _valid_date_spec(spec: str) -> bool:
            if parse_yyyymmdd_flex(spec):
                return True
            return bool(get_date_relative_to(spec, date.today()))  # noqa: DTZ011

        if not select_config:
            config_warning("Select contains no conditions.", workbook.config, config_keys)

        if select_config.get("date_from") or select_config.get("date_to"):
            for k in ("date_from", "date_to"):
                v = select_config.get(k)
                if v and not _valid_date_spec(v):
                    raise MnyXlsConfigError(f"'{v}': Invalid date spec.", workbook.config, (*config_keys, k))

            if select_config.get("yyyy"):
                config_warning("`date_from` and `date_to` are ignored when `yyyy` is set.", workbook.config, config_keys)

        # Sanity check select criteria
        for select_key in select_config:
            values = get_select_values(select_key, select_config)
            if not values:
                raise MnyXlsConfigError("At least one value is required.", workbook.config, (*config_keys, select_key))

    def add_worksheet(self, worksheet: MoneyWorksheet) -> None:
        """Add worksheet to workbook.

        Args:
            worksheet (MoneyWorksheet): Worksheet to add.

        Raises:
            MnyXlsConfigError: Invalid configuration.
        """
        assert worksheet.workbook == self
        self.worksheets.append(worksheet)

    def create_worksheets(self, conn: sqlite3.Connection) -> None:
        """Add worksheets to workbook.

        A single worksheet configuration block can be expanded into multiple
        worksheets based on configuration scenarios such as `foreach` directives.

        Args:
            conn (sqlite3.Connection): SQLite connection.

        Raises:
            MnyXlsConfigError: Invalid configuration.
        """
        for worksheet_cls, sheet_name_or_key, worksheet_config in self.worksheet_specs:
            if worksheet_cls.expand_config(self, conn, sheet_name_or_key, worksheet_config) == 0:
                self.add_worksheet(worksheet_cls(self, sheet_name_or_key, worksheet_config))

        if len(self.worksheets) > WORKBOOK_MAX_SHEETS:
            raise MnyXlsConfigError(f"Too many worksheets; limit is {WORKBOOK_MAX_SHEETS}", self.config, "workbook")

    def write_workbook(self, conn: sqlite3.Connection) -> None:
        """Write workbook to file.

        Args:
            conn (sqlite3.Connection): SQLite connection.
            xls_file (Path): Path to workbook.

        Raises:
            MnyXlsConfigError: Invalid configuration.
        """
        template_wb = None
        writer_opts = {}

        assert self.xls_file is not None

        logger.debug(f"Workbook path: '{self.xls_file}'")
        logger.debug(f"Workbook template path: '{self.xls_template}'")

        if self.xls_template is not None:
            try:
                template_wb = load_workbook(self.xls_template)
            except FileNotFoundError:
                logger.debug(f"{self.xls_template.name}: No such file; creating new workbook")
            else:
                logger.debug(f"Copied workbook template to '{self.xls_file.name}'")

                # Copy template to output file
                template_wb.save(self.xls_file)

                # Keeping sheets defined in the template
                writer_opts["mode"] = "a"
                writer_opts["if_sheet_exists"] = "replace"

                # Existing sheets can be replaced.
                self.can_use_existing_sheet = dict.fromkeys(template_wb.sheetnames, True)

        logger.debug(f"Writing workbook '{self.xls_file.name}'...")

        with pd.ExcelWriter(self.xls_file, engine="openpyxl", **writer_opts) as writer:
            # ExcelWriter aborts ungracefully if an exception is raised,
            # mostly because it requires a sheet be written, otherwise it raises
            # > IndexError("At least one sheet must be visible").
            # We work around this by creating a temporary empty sheet.
            # This is removed after all sheets are written.
            pd.DataFrame().to_excel(writer, sheet_name=TEMPORARY_SHEET_NAME, index=False)

            empty_sheets: list[int] = []

            for i, worksheet in enumerate(self.worksheets):
                worksheet.prepare_write_sheet(conn, writer)

                if worksheet.write_sheet(conn, writer) is None:
                    empty_sheets.append(i)
                    continue

                worksheet.format_sheet()

            # Worksheets that are not written are not included in `autofit_columns`.
            for i in sorted(empty_sheets, reverse=True):
                self.empty_worksheets.append(self.worksheets[i])
                del self.worksheets[i]

            self.finalize_workbook()

            if len(writer.sheets) > 1:
                del writer.book[TEMPORARY_SHEET_NAME]
            else:
                logger.warning("No worksheets were generated.")

    def finalize_workbook(self) -> None:
        """Called after all sheets are written to the workbook."""
        self.autofit_columns()

        for worksheet in self.worksheets:
            worksheet.finalize_sheet()

        logger.debug("Workbook complete!")

    def autofit_columns(self) -> None:
        """Apply consistent column widths across sheets of same type based on maximum column value lengths."""
        autofit_workbook = bool(self.workbook_config.get("autofit", True))

        for sheet_type, worksheets in self._get_worksheets_by_type().items():
            sheet_type_column_widths: dict[DataFrameKeyT, int] = {}  # column: width

            for worksheet in worksheets:
                sheet_column_widths: dict[DataFrameKeyT, int] = worksheet.column_value_maxlen
                for column, width in sheet_column_widths.items():
                    sheet_type_column_widths[column] = int(max(sheet_type_column_widths.get(column, 0), width))

            if False:
                logger.debug(f"Setting column widths across '{sheet_type}' sheets: {sheet_type_column_widths}")

            for worksheet in worksheets:
                autofit_worksheet = worksheet.worksheet_config.get("autofit")
                autofit_worksheet = autofit_workbook if autofit_worksheet is None else bool(autofit_worksheet)

                if autofit_worksheet:
                    worksheet.autofit_columns(width_hints=sheet_type_column_widths)

    ######################################################################
    # Helper methods

    def _get_worksheets_by_type(self) -> dict[str, Sequence[MoneyWorksheet]]:
        """Return a dictionary of worksheets by type.

        Returns:
            dict[str, Sequence[MoneyWorksheet]]
        """
        worksheets_by_type: dict[str, list[MoneyWorksheet]] = defaultdict(list)

        for worksheet in self.worksheets:
            worksheets_by_type[worksheet.SHEET_TYPE].append(worksheet)

        return dict(worksheets_by_type)


######################################################################
# Module public


def gather_workbook(ctx: click.Context, xls_file: Path, config: MainConfigFileT) -> MoneyWorkbook:  # noqa: ARG001
    """Create workbook object from configuration file directives.

    Args:
        ctx (click.Context): Click context.
        xls_file (Path): Path to workbook.
        config (ConfigFileT): User configuration directives.

    Returns:
        MoneyWorkbook

    Raises:
        MnyXlsConfigError: Configuration error.
    """
    xls_config = None
    xls_config_file = config.get("xls_config")

    # Workbook configuration can be specified in a separate file.
    if xls_config_file and xls_config_file.is_file():
        try:
            xls_config = read_config_file(xls_config_file)
        except MnyXlsConfigError as err:
            raise MnyXlsConfigError(f"Failed to read workbook config file '{xls_config_file}': {err}") from err
        else:
            xls_config = typing.cast("WorkbookConfigFileT", xls_config)
            xls_config["_config_file"] = xls_config_file  # for error reporting

    workbook = MoneyWorkbook(xls_file, xls_config, config)
    workbook.validate_config()

    return workbook

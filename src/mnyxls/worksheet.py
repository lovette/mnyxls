from __future__ import annotations

import logging
import typing
from abc import ABC, abstractmethod
from datetime import date
from functools import cached_property
from types import MappingProxyType
from typing import TYPE_CHECKING, Any

import numpy as np
import pandas as pd
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import DEFAULT_BORDER
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pandas.io.formats.excel import ExcelFormatter  # type: ignore[reportMissingImport]

from .configtypes import WorksheetConfigSelectT, WorksheetConfigT
from .jinja import render_template_str
from .shared import (
    MnyXlsConfigError,
    truncate_w_ellipsis,
    validate_config_typed_dict,
)

if TYPE_CHECKING:
    import sqlite3
    from collections.abc import Generator, Mapping, Sequence

    from .workbook import MoneyWorkbook


WORKSHEET_NAME_LEN = 31  # Excel sheet names limited to 31 chars
WORKSHEET_DEFAULT_FORMAT_DATE = "m/d/yy"  # Excel format string
WORKSHEET_DEFAULT_FORMAT_AMOUNT = "$#,##0.00_);[Red]($#,##0.00)"
WORKSHEET_NAME_ILLEGAL_CHARS = tuple(r"*/\?*[]")  # Excel sheet names cannot contain these characters
WORKSHEET_CURRENCY_PADDING = 3  # Add space for the '$()' that Excel includes for currency values
WORKSHEET_COLWIDTH_MAX = 45  # Arbitrary
WORKSHEET_COLUMNS_MAX = 500  # Ensure a reasonable number (pivot tables can get out of hand!)

FORMAT_FILL_NONE = PatternFill(fill_type=None)
FORMAT_BORDER_NONE = DEFAULT_BORDER
FORMAT_ALIGN_LEFT = Alignment(horizontal="left", vertical="top")
FORMAT_ALIGN_CENTER = Alignment(horizontal="center", vertical="top")
FORMAT_NORMAL_FONT = Font(bold=False)
FORMAT_BOLD_FONT = Font(bold=True)

FRIENDLY_COLUMN_NAMES = MappingProxyType(
    {
        "AccountCategory": "Account Category",
        "AccountClassification": "Account Classification",
        "AccountLimit": "Account Limit",
        "AccountNumber": "Account Number",
        "BankName": "Bank Name",
        "ClosedDate": "Closed Date",
        "DateYYYY": "Year",
        "DateYYYYMM": "Month",
        "OpenedDate": "Opened Date",
        "OpeningBalance": "Opening Balance",
        "TxnClass": "Txn Class",
        "TxnDateMax": "Txn Date Max",
        "TxnDateMin": "Txn Date Min",
        "TxnType": "Txn Type",
        "XferAccount": "Xfer Account",
    }
)

ExcelWriterWorksheetT = Worksheet  # to_excel does not define a "worksheet" type
DataFrameKeyT = str | tuple[str, ...]  # Hashable keys used to index DataFrame columns (df[col])
HeaderToLetterMapT = dict[DataFrameKeyT, str]
WorksheetCellValueT = str | int | float | date | Any | None  # Any=ArrayFormula, CellRichText, etc.


# Classes derived from `MoneyWorksheet` use `register_sheet_type` to register themselves here.
_WORKSHEET_TYPE_CLS_MAP: dict[str, type[MoneyWorksheet]] = {}

# `pd.to_excel()` uses this property to format index and column header cells.
# This is a global setting for all worksheets.
#
# The default is:
# > {
# >     "font": {"bold": True},
# >     "borders": {
# >         "top": "thin",
# >         "right": "thin",
# >         "bottom": "thin",
# >         "left": "thin",
# >     },
# >     "alignment": {"horizontal": "center", "vertical": "top"},
# > }
ExcelFormatter.header_style = {
    "font": {"bold": True},
    "alignment": {"horizontal": "center", "vertical": "top"},
}

# Use a named logger instead of root logger
logger = logging.getLogger("mnyxls")


######################################################################
# Worksheet base class


class MoneyWorksheet(ABC):
    """Worksheet base class."""

    # Keep column widths within reasonable bounds
    WORKSHEET_COLWIDTH_BOUNDS = MappingProxyType({})  # derived class sets this

    SHEET_TYPE: str = "Undefined"  # derived class sets this

    def __init__(self, workbook: MoneyWorkbook, sheet_name: str, config: WorksheetConfigT) -> None:
        """Constructor.

        Args:
            workbook (MoneyWorkbook): Workbook.
            sheet_name (str): Proposed sheet name/label; actual name is determined when sheet is written.
            config (WorksheetConfigT): User configuration options.
        """
        assert workbook is not None
        assert sheet_name is not None
        assert len(sheet_name) > 0

        self.workbook = workbook
        self.config_key = config.get("_config_key")
        self.sheet_name = sheet_name  # may be renamed by `write_sheet`
        self.worksheet_config = config
        self.df_worksheet: pd.DataFrame | None = None
        self.pyxl_worksheet: ExcelWriterWorksheetT | None = None  # Set by `to_excel`
        self.preserve_column_headers: Sequence[str] | None = None
        self.date_cols: Sequence[DataFrameKeyT] | None = None  # List of columns of type `date`
        self.currency_cols: Sequence[DataFrameKeyT] | None = None  # List of columns of type `CurrencyDecimal`
        self.index_to_excel = False

        # 1-based bottommost row and rightmost column that is to be frozen.
        # Freeze the first row by default
        self.freeze_panes: tuple[int, int] | None = (1, 1)

        assert self.config_key
        self.config_keys = ("workbook", "worksheets", self.config_key)

        date_number_format = self.get_config_value("date_number_format", WORKSHEET_DEFAULT_FORMAT_DATE)
        assert isinstance(date_number_format, str)
        self.date_number_format: str = date_number_format

        amount_number_format = self.get_config_value("amount_number_format", WORKSHEET_DEFAULT_FORMAT_AMOUNT)
        assert isinstance(amount_number_format, str)
        self.amount_number_format: str = amount_number_format

    def __repr__(self) -> str:
        """Return a string representation of the object for debugging purposes.

        Returns:
            str
        """
        return f"{self.__class__.__name__}('{self.sheet_name}')"

    ######################################################################
    # Instance properties

    @cached_property
    def header_levels(self) -> int:
        """Number of header levels.

        This is the number of rows in the header based on the DataFrame returned by `get_sheet_data`.
        It will be greater than 1 if the DataFrame has a MultiIndex.
        Index columns may have more header rows than data columns.
        The number of data column header rows is `self.df_worksheet.columns.nlevels`.
        Fails until `get_sheet_data` is complete.

        Returns:
            int
        """
        assert self.df_worksheet is not None

        header_levels_ = self.df_worksheet.columns.nlevels

        if header_levels_ > 1:
            # When columns are MultiIndex the worksheet has an additional header row for the index
            header_levels_ += 1

        return header_levels_

    @cached_property
    def index_levels(self) -> int:
        """Number of index levels.

        This is the number of columns in the index based on the DataFrame returned by `get_sheet_data`.
        It will be greater than 1 if the DataFrame has a MultiIndex.
        Fails until `get_sheet_data` is complete.

        Returns:
            int
        """
        assert self.df_worksheet is not None
        return self.df_worksheet.index.nlevels

    @cached_property
    def column_widths(self) -> dict[str, int]:
        """Current worksheet column widths by header value.

        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.

        Returns:
            dict[str, int]
        """
        assert self.pyxl_worksheet is not None

        widths: dict[str, int] = {}

        for cell in self.iter_header_cells(1):
            column_letter, row_idx = coordinate_from_string(cell.coordinate)  # handles merged cells
            widths[str(cell.value)] = int(self.pyxl_worksheet.column_dimensions[column_letter].width)

        return widths

    @cached_property
    def column_value_maxlen(self) -> dict[DataFrameKeyT, int]:
        """Column maximum value string lengths by header value(s).

        Dictionary keys map to DataFrame column indexes.
        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.

        Returns:
            dict[str, int]
        """
        assert self.pyxl_worksheet is not None

        padding = 0
        column_value_maxlen: dict[DataFrameKeyT, int] = {}

        for col_df_idx, letter in self.header_to_letter_map.items():
            maxlen = 0

            for col_cell in self.pyxl_worksheet[letter]:
                assert isinstance(col_cell, Cell | MergedCell)
                assert col_cell.row is not None

                if col_cell.value is None:
                    continue  # merged cell or pivot header

                if self.currency_cols and col_df_idx in self.currency_cols and self.header_levels < col_cell.row:
                    padding = WORKSHEET_CURRENCY_PADDING

                maxlen = max(len(str(col_cell.value)) + padding, maxlen)

            column_value_maxlen[col_df_idx] = int(maxlen)

        return column_value_maxlen

    @cached_property
    def header_to_letter_map(self) -> HeaderToLetterMapT:
        """Return a dictionary that maps header column values to worksheet column letters.

        Each DataFrame column (`self.df_worksheet.index.columns`) is mapped to a worksheet column.
        Each DataFrame index name (`self.df_worksheet.index.names`) is mapped to a worksheet column.
        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.

        Returns:
            dict[header values, letter]
        """
        assert self.pyxl_worksheet is not None
        assert self.df_worksheet is not None

        header_map: HeaderToLetterMapT = {}

        # A MultiIndex that looks like this, where X, Y are index groups
        # and B, M are column groups. This matches the MultiIndex column indexes
        # used in the DataFrame.
        #
        #   | B | C
        #   | M | N
        # X | Y |
        #
        # Will be represented as:
        #
        # > [[None, 'B', 'C'], [None, 'M', 'N'], ['X', 'Y', None]]
        # > [['', '', 'X'], ['B', 'M', 'Y'], ['C', 'N', '']]
        # > ['X', 'Y', ('C', 'N')]

        header_rows_values = [self._get_row_values(row_cells) for row_cells in self.pyxl_worksheet.iter_rows(max_row=self.header_levels)]
        column_indexes = list(zip(*header_rows_values, strict=True))

        for idx_col, header_values in enumerate(column_indexes, start=1):
            if len(header_values) > 1:
                header_values = header_values[-1] if idx_col <= self.index_levels else header_values[:-1]  # noqa: PLW2901

            header_map[header_values[0] if len(header_values) == 1 else header_values] = get_column_letter(idx_col)

        # Sanity check that we can map all DataFrame index and column names
        for col in self.df_worksheet.index.names:
            if col is not None:  # RangeIndex values are unnamed integers
                assert header_map.get(col) is not None
        for col in self.df_worksheet.columns:
            assert header_map.get(col) is not None

        return header_map

    @cached_property
    def worksheet_title(self) -> str:
        """Worksheet title.

        Fails until `to_excel` is complete.

        Returns:
            str
        """
        assert self.pyxl_worksheet is not None
        return self.pyxl_worksheet.title

    @property
    def in_workbook(self) -> bool:
        """Return True if worksheet was written to workbook.

        Returns:
            bool
        """
        return self.pyxl_worksheet is not None

    ######################################################################
    # Class methods

    @classmethod
    def register_sheet_type(cls) -> None:
        """Register worksheet class so it can be referenced in the configuration file."""
        assert cls.SHEET_TYPE != ""
        _WORKSHEET_TYPE_CLS_MAP[cls.SHEET_TYPE] = cls

    @classmethod
    def get_worksheet_class(cls, worksheet_type: str) -> type[MoneyWorksheet] | None:
        """Get worksheet class by type.

        This is used to create a worksheet from the configuration file.

        Args:
            worksheet_type (str): Worksheet type.

        Returns:
            type[MoneyWorksheet] | None
        """
        assert len(_WORKSHEET_TYPE_CLS_MAP) > 0
        return _WORKSHEET_TYPE_CLS_MAP.get(worksheet_type)

    @classmethod
    def validate_config(cls, workbook: MoneyWorkbook, sheet_config_key: str, worksheet_config: WorksheetConfigT) -> None:
        """Validate worksheet configuration.

        Args:
            workbook (MoneyWorkbook): Workbook.
            sheet_config_key (str): Sheet name.
            worksheet_config (WorksheetConfigT): Worksheet configuration.

        Raises:
            MnyXlsConfigError: Invalid configuration.
        """
        assert workbook is not None

        config_keys = ("workbook", "worksheets", sheet_config_key)

        if not worksheet_config:
            raise MnyXlsConfigError("A worksheet must be defined.", workbook.config, config_keys)

        validate_config_typed_dict(worksheet_config, WorksheetConfigT, workbook.config, config_keys)

        config_select = worksheet_config.get("select")
        if config_select is not None:
            cls.validate_config_select(workbook, config_select, (*config_keys, "select"))

    @classmethod
    def validate_config_select(cls, workbook: MoneyWorkbook, select_config: WorksheetConfigSelectT, config_keys: str | Sequence[str]) -> None:
        """Validate worksheet configuration `select` directive.

        Args:
            workbook (MoneyWorkbook): Workbook.
            select_config (WorksheetConfigSelectT): Select configuration.
            config_keys (str | Sequence[str]): Configuration keys for reporting errors.

        Raises:
            MnyXlsConfigError: Invalid configuration.
        """
        workbook.validate_config_select(workbook, select_config, config_keys)

    @classmethod
    def expand_config(
        cls,
        workbook: MoneyWorkbook,
        conn: sqlite3.Connection,
        worksheet_key: str,
        worksheet_base_config: WorksheetConfigT,
    ) -> int:
        """Create a worksheet for each configuration scenario and add to workbook.

        Args:
            workbook (MoneyWorkbook): Workbook.
            conn (sqlite3.Connection): SQLite connection.
            worksheet_key (str): Worksheet configuration key (typically the sheet name.)
            worksheet_base_config (WorksheetConfigT): Worksheet configuration.

        Returns:
            int: Number of worksheets added to workbook.

        Raises:
            MnyXlsConfigError: Configuration error.
        """
        if "foreach" not in worksheet_base_config:
            return 0  # No expansions made

        # Expand 'foreach' to create individual worksheets
        return cls.expand_foreach(workbook, conn, worksheet_key, worksheet_base_config)

    @classmethod
    def expand_foreach(
        cls,
        workbook: MoneyWorkbook,
        conn: sqlite3.Connection,
        worksheet_key: str,
        worksheet_base_config: WorksheetConfigT,
    ) -> int:
        """Create a worksheet for each 'foreach' value and add to workbook.

        Args:
            workbook (MoneyWorkbook): Workbook.
            conn (sqlite3.Connection): SQLite connection.
            worksheet_key (str): Worksheet configuration key (typically the sheet name.)
            worksheet_base_config (WorksheetConfigT): Worksheet configuration.

        Returns:
            int: Number of worksheets added to workbook.

        Raises:
            MnyXlsConfigError: Configuration error.
        """
        config_keys = ("workbook", "worksheets", worksheet_key)
        config_foreach = worksheet_base_config.get("foreach")

        assert config_foreach is not None

        # Lookup a classmethod named to handle this `foreach` directive.
        cls_fn_expand_foreach = getattr(cls, f"expand_foreach_{config_foreach}", None)

        if cls_fn_expand_foreach is None:
            raise MnyXlsConfigError("Unrecognized option.", workbook.config, (*config_keys, "foreach", config_foreach))

        return cls_fn_expand_foreach(workbook, conn, worksheet_key, worksheet_base_config)

    @classmethod
    def copy_worksheet_config(cls, c: WorksheetConfigT) -> WorksheetConfigT:
        """Create a *shallow* copy of the worksheet configuration with a *modifiable copy* of the `select` directive.

        Args:
            c (WorksheetConfigT): Worksheet configuration.

        Returns:
            WorksheetConfigT
        """
        config_copy = c.copy()

        if "select" in config_copy:
            config_copy["select"] = typing.cast("WorksheetConfigSelectT", dict(config_copy["select"]))
        else:
            config_copy["select"] = {}

        return config_copy

    @classmethod
    def render_sheet_name(cls, sheet_name: str, template_vars: Mapping[str, Any], default_name: str) -> str:
        """Render a worksheet name using Jinja2.

        Args:
            sheet_name (str): Template string to render.
            template_vars (Mapping[str, Any]): Template variables.
            default_name (str): Default name if `sheet_name` is not a template.

        Returns:
            str
        """
        if "{{" not in sheet_name:
            return default_name

        sheet_name = render_template_str(sheet_name, template_vars)

        return sheet_name or default_name

    ######################################################################
    # Instance methods

    def get_config_value(self, key: str, default: str | int | bool | None) -> str | int | bool | Sequence | Mapping | None:
        """Get a worksheet configuration value.

        Worksheet configuration values override workbook and top-level configuration values.

        Args:
            key (str): Configuration key.
            default (str): Default value if key is not set.

        Returns:
            str | int | bool | Sequence | Mapping | None
        """
        assert self.workbook is not None
        assert self.worksheet_config is not None

        return self.worksheet_config.get(key, self.workbook.get_config_value(key, default))

    def prepare_write_sheet(self, conn: sqlite3.Connection, writer: pd.ExcelWriter) -> None:
        """Prepare to write worksheet to workbook.

        Invoked prior to `write_sheet`.

        Args:
            conn (sqlite3.Connection): SQLite connection.
            writer (pd.ExcelWriter): Excel writer.

        Raises:
            MnyXlsConfigError
        """
        assert conn is not None
        assert writer is not None

    def write_sheet(self, conn: sqlite3.Connection, writer: pd.ExcelWriter) -> ExcelWriterWorksheetT | None:
        """Write worksheet to workbook.

        CurrencyDecimal columns listed in `self.currency_cols` will be saved as floats to
        avoid Excel marking cells with "Number Stored as Text".

        ```
        Workflow is:
        1. prepare_write_sheet
        2. write_sheet
        3. ...get_sheet_data (abstractmethod)
        4. ...prepare_to_excel
        5. ...to_excel
        6. format_sheet
        7. autofit_columns
        8. finalize_sheet
        ```

        Args:
            conn (sqlite3.Connection): SQLite connection.
            writer (pd.ExcelWriter): Excel writer.

        Returns:
            ExcelWriterWorksheetT: Excel worksheet or None if sheet was not added to workbook.
        """
        assert self.df_worksheet is None
        assert self.pyxl_worksheet is None

        writer_sheet_names_uc = {sheet_name.upper() for sheet_name in writer.sheets}

        safe_sheet_name = self._safe_sheet_name(self.sheet_name, writer_sheet_names_uc)

        if safe_sheet_name in writer.sheets:
            # Capture column headers in the existing sheet so we can preserve the order later.
            ws: Worksheet = writer.sheets[safe_sheet_name]
            self.preserve_column_headers = tuple([str(cells[0].value) for cells in ws.iter_cols(max_row=1)])

        self.df_worksheet = self.get_sheet_data(conn)
        df_worksheet = self.df_worksheet

        skipempty = self.worksheet_config.get("skipempty", True)

        if df_worksheet.empty and skipempty:
            logger.debug(f"... worksheet '{safe_sheet_name}' has no data, skipping")
            return None  # do not create empty worksheet

        if len(df_worksheet.columns) > WORKSHEET_COLUMNS_MAX:
            logger.warning(f"Worksheet '{safe_sheet_name}' has {len(df_worksheet.columns)} columns (the limit is {WORKSHEET_COLUMNS_MAX}), skipping")
            return None

        logger.debug(f"... worksheet '{safe_sheet_name}' with {df_worksheet.shape[0]} rows")

        df_worksheet = self.prepare_to_excel()

        return self.to_excel(writer, safe_sheet_name, df_worksheet)

    @abstractmethod
    def get_sheet_data(self, conn: sqlite3.Connection) -> pd.DataFrame:
        """Query database and return data to write to worksheet.

        Args:
            conn (sqlite3.Connection): SQLite connection.
        """
        raise NotImplementedError

    def format_sheet(self) -> None:
        """Called after the sheet is written to the workbook."""
        assert self.pyxl_worksheet is not None

    def finalize_sheet(self) -> None:
        """Called after all sheets are written to the workbook."""
        assert self.pyxl_worksheet is not None

        # Replace header values with more user-friendly values
        # This is done just before the workbook is save so that worksheet columns names
        # match dataframe columns until the very end.
        for cell in self.iter_header_cells():
            if isinstance(cell.value, str) and cell.value in FRIENDLY_COLUMN_NAMES:
                cell.value = FRIENDLY_COLUMN_NAMES[cell.value]

    def prepare_to_excel(self) -> pd.DataFrame:
        """Prepare to write worksheet to workbook.

        Invoked prior to `to_excel`.
        `self.df_worksheet` is set to the result of `get_sheet_data`.

        Returns:
            pd.DataFrame: DataFrame to write to workbook.
        """
        assert self.df_worksheet is not None

        df_worksheet = self.df_worksheet

        # RangeIndex values are unnamed integers and don't need to be included.
        # Pivot table indexes must be included.
        self.index_to_excel = not isinstance(df_worksheet.index, pd.RangeIndex)

        if self.currency_cols:
            df_worksheet = df_worksheet.copy()  # don't modify original DataFrame
            for col in self.currency_cols:
                # Write currency as a float to avoid Excel marking cells with "Number Stored as Text".
                df_worksheet[col] = df_worksheet[col].apply(lambda x: x.nofmt() if pd.notna(x) else np.nan).astype(float)

        return df_worksheet

    def to_excel(self, writer: pd.ExcelWriter, sheet_name: str, df_worksheet: pd.DataFrame) -> ExcelWriterWorksheetT | None:
        """Write data to workbook sheet.

        Args:
            writer (pd.ExcelWriter): Excel writer.
            sheet_name (str): Worksheet name.
            df_worksheet (pd.DataFrame): Worksheet data.

        Returns:
            ExcelWriterWorksheetT: Excel worksheet or None if sheet was not added to workbook.
        """
        # Create worksheet
        df_worksheet.to_excel(
            writer,
            sheet_name=sheet_name,
            float_format="%.2f",
            freeze_panes=self.freeze_panes,
            index=self.index_to_excel,
        )

        if sheet_name not in writer.sheets:
            # The sheet name was changed by `to_excel` to avoid duplicates.
            last_sheet = list(writer.sheets.keys())[-1]
            logger.warning(f"Worksheet '{sheet_name}' unexpectedly renamed to '{last_sheet}'.")
            sheet_name = last_sheet

        self.pyxl_worksheet = writer.sheets[sheet_name]
        assert self.pyxl_worksheet is not None

        return self.pyxl_worksheet

    def set_column_number_format(  # noqa: PLR0913
        self,
        cell_format: str,
        header: DataFrameKeyT | None = None,
        letter: str | None = None,
        idx: int | None = None,
        header_levels: int | None = None,
        raise_missing_header: bool = True,
    ) -> None:
        """Set Excel number format for non-header cells in column with given index, header or letter.

        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.

        Args:
            cell_format (str): Cell `number_format`.
            header (DataFrameIndexHashableT): Column header value.
            letter (str): Column letter.
            idx (str): Column index; 1-based.
            header_levels (int): Header levels; 1 if header includes only the first row.
            skipmerged (bool): True to skip merged cells.
            raise_missing_header (bool): True to raise `ValueError` if header value cannot be found.
        """
        # Set format for non-header cells in a column (idx=0 is the header row.)
        for row_cell in self.iter_data_rows(
            header=header,
            letter=letter,
            idx=idx,
            header_levels=header_levels,
            skipmerged=True,
            raise_missing_header=raise_missing_header,
        ):
            row_cell.number_format = cell_format

    def autofit_columns(self, width_hints: dict[DataFrameKeyT, int] | None = None) -> None:
        """Autofit column widths based on maximum value string lengths.

        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.
        Column headers that do not exist in the worksheet are ignored.
        Column widths are set to the maximum value string length by default.

        Args:
            width_hints (dict[str, int]): Column width hints by column header.
                Sheet is not required to honor these hints.
        """
        # Setting 'auto_size' attribute does not (and will never) work :/
        # https://stackoverflow.com/questions/60248319/how-to-set-column-width-to-bestfit-in-openpyxl
        # https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1275
        # > self.pyxl_worksheet.column_dimensions[col[0].column_letter].auto_size = True

        column_widths = self.column_value_maxlen if width_hints is None else width_hints

        self.set_column_widths(column_widths)

    def set_column_widths(self, col_widths: dict[DataFrameKeyT, int]) -> None:
        """Set column widths using column header values.

        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.
        Column headers that do not exist in the worksheet are ignored.

        Args:
            col_widths (dict[str, int]): Column widths by column header.
        """
        assert self.pyxl_worksheet is not None
        assert len(self.WORKSHEET_COLWIDTH_BOUNDS) > 0

        column_letters = self.header_to_letter_map

        for col_df_idx, width in col_widths.items():
            col_letter = column_letters.get(col_df_idx)
            if col_letter is not None:
                if col_df_idx in self.WORKSHEET_COLWIDTH_BOUNDS:
                    min_width, max_width = self.WORKSHEET_COLWIDTH_BOUNDS[col_df_idx]
                    width = max(min_width, width)  # noqa: PLW2901
                    width = min(max_width, max(min_width, width))  # noqa: PLW2901

                self.pyxl_worksheet.column_dimensions[col_letter].width = width

    def iter_header_cells(
        self,
        level: int | None = None,
        max_level: int | None = None,
        skipmerged: bool = True,
        skipnovalue: bool = True,
    ) -> Generator[Cell | MergedCell, None, None]:
        """Produces cells from the worksheet header, column by column, row by row.

        Iterates row-by-row in multi-row headers if `max_level` is greater than `level`.
        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.

        Args:
            level (int): Header level; 1 is the first row.
            max_level (int): Number of levels to iterate; Default is `header_levels`.
            skipmerged (bool): True to skip merged cells. Keep in mind: The `value` of merged cells is None.
            skipnovalue (bool): True to skip cells with value `None` such as those created for a MultiIndex dataframe.

        Yields:
            Cell | MergedCell
        """
        assert self.pyxl_worksheet is not None

        if level is None and max_level is None:
            level = 1
            max_level = self.header_levels
        elif max_level is None:
            max_level = level
        elif level is None:
            level = 1

        for row_cells in self.pyxl_worksheet.iter_rows(min_row=level, max_row=max_level):
            for cell in row_cells:
                if isinstance(cell, MergedCell):
                    if skipmerged:
                        continue
                else:
                    assert isinstance(cell, Cell)
                    if cell.value is None and skipnovalue:
                        continue

                yield cell

    def iter_header_rows(  # noqa: PLR0913
        self,
        header: DataFrameKeyT | None = None,
        letter: str | None = None,
        idx: int | None = None,
        cell: Cell | MergedCell | None = None,
        header_levels: int | None = None,
        skipmerged: bool = True,
        raise_missing_header: bool = True,
    ) -> Generator[Cell | MergedCell, None, None]:
        """Produces cells from a worksheet header column, by row.

        Column can be identified by an index, header or letter.
        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.

        Args:
            header (DataFrameIndexHashableT): Column header value; None if `letter` or `idx` is used.
            letter (str): Column letter (A, B, ...); None if `header` or `idx` is used.
            idx (str): Column index; (1=A, 2=B, ...); None if `header` or `letter` is used.
            cell (Cell | MergedCell): Cell reference; None if `header`, `letter` or `idx` is used.
            header_levels (int): Number of levels to iterate; Default is `header_levels`.
            skipmerged (bool): True to skip merged cells. Keep in mind: The `value` of merged cells is None.
            raise_missing_header (bool): True to raise `ValueError` if header value cannot be found.

        Yields:
            Cell | MergedCell

        See Also:
            iter_data_rows
        """
        assert self.pyxl_worksheet is not None

        idx = self._get_header_letter_idx(
            header=header,
            letter=letter,
            idx=idx,
            cell=cell,
            raise_missing_header=raise_missing_header,
        )

        if idx is None:
            raise ValueError("A column reference index, header or letter must be set.")

        if header_levels is None:
            header_levels = self.header_levels

        for row_cells in self.pyxl_worksheet.iter_rows(min_col=idx, max_col=idx, max_row=header_levels):
            assert len(row_cells) == 1
            row_cell = row_cells[0]
            assert isinstance(row_cell, Cell | MergedCell)
            if isinstance(row_cell, MergedCell) and skipmerged:
                continue
            yield row_cell

    def iter_data_rows(  # noqa: PLR0913
        self,
        header: DataFrameKeyT | None = None,
        letter: str | None = None,
        idx: int | None = None,
        cell: Cell | MergedCell | None = None,
        header_levels: int | None = None,
        skipmerged: bool = True,
        raise_missing_header: bool = True,
    ) -> Generator[Cell | MergedCell, None, None]:
        """Produces cells from a worksheet data column, by row.

        Column can be identified by an index, header or letter.
        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.

        Args:
            header (DataFrameIndexHashableT): Column header value; None if `letter` or `idx` is used.
            letter (str): Column letter (A, B, ...); None if `header` or `idx` is used.
            idx (str): Column index; (1=A, 2=B, ...); None if `header` or `letter` is used.
            cell (Cell | MergedCell): Cell reference; None if `header`, `letter` or `idx` is used.
            header_levels (int): Number of levels to iterate; Default is `header_levels`.
            skipmerged (bool): True to skip merged cells. Keep in mind: The `value` of merged cells is None.
            raise_missing_header (bool): True to raise `ValueError` if header value cannot be found.

        Yields:
            Cell | MergedCell

        See Also:
            iter_header_rows
        """
        assert self.pyxl_worksheet is not None

        idx = self._get_header_letter_idx(
            header=header,
            letter=letter,
            idx=idx,
            cell=cell,
            raise_missing_header=raise_missing_header,
        )

        if idx is None:
            raise ValueError("A column reference index, header or letter must be set.")

        if header_levels is None:
            header_levels = self.header_levels

        for row_cells in self.pyxl_worksheet.iter_rows(min_col=idx, max_col=idx, min_row=header_levels + 1):
            assert len(row_cells) == 1
            row_cell = row_cells[0]
            assert isinstance(row_cell, Cell | MergedCell)
            if isinstance(row_cell, MergedCell) and skipmerged:
                continue
            yield row_cell

    def iter_index_cols(
        self,
        idx: int | None = None,
        cell: Cell | MergedCell | None = None,
        skipmerged: bool = True,
    ) -> Generator[Cell | MergedCell, None, None]:
        """Produces cells from worksheet index columns, by column.

        The first `index_levels` columns are index columns.

        Args:
            idx (int): Row index, 1-based; None if `cell` is used. Defaults to 1.
            cell (Cell | MergedCell): Row cell reference; None if `idx` is used.
            skipmerged (bool): True to skip merged cells. Keep in mind: The `value` of merged cells is None.

        Yields:
            Cell | MergedCell

        See Also:
            iter_data_cols
        """
        assert self.pyxl_worksheet is not None

        if idx is None and cell is not None:
            assert cell.row is not None
            idx = cell.row
        elif idx is None:
            idx = 1

        for col_cells in self.pyxl_worksheet.iter_cols(min_row=idx, max_row=idx, max_col=self.index_levels):
            assert len(col_cells) == 1
            col_cell = col_cells[0]
            assert isinstance(col_cell, Cell | MergedCell)
            if isinstance(col_cell, MergedCell) and skipmerged:
                continue
            yield col_cell

    def iter_data_cols(
        self,
        idx: int | None = None,
        cell: Cell | MergedCell | None = None,
        skipmerged: bool = True,
    ) -> Generator[Cell | MergedCell, None, None]:
        """Produces cells from worksheet data columns, by column.

        All columns after the first `index_levels` columns are data columns.

        Args:
            idx (int): Row index, 1-based; None if `cell` is used. Defaults to first data row.
            cell (Cell | MergedCell): Row cell reference; None if `idx` is used.
            skipmerged (bool): True to skip merged cells. Keep in mind: The `value` of merged cells is None.

        Yields:
            Cell | MergedCell

        See Also:
            iter_index_cols
        """
        assert self.pyxl_worksheet is not None

        if idx is None and cell is not None:
            assert cell.row is not None
            idx = cell.row
        elif idx is None:
            idx = self.header_levels + 1

        for col_cells in self.pyxl_worksheet.iter_cols(min_row=idx, max_row=idx, min_col=self.index_levels + 1):
            assert len(col_cells) == 1
            col_cell = col_cells[0]
            assert isinstance(col_cell, Cell | MergedCell)
            if isinstance(col_cell, MergedCell) and skipmerged:
                continue
            yield col_cell

    ######################################################################
    # Helper methods

    def _safe_sheet_name(self, requested_name: str, writer_sheet_names_uc: set[str]) -> str:  # noqa: C901
        """Return a valid and potentially unique worksheet name.

        Args:
            writer_sheet_names_uc (set[str]): Set of existing worksheet names, uppercased.
            requested_name (str): Requested worksheet name.

        Returns:
            str
        """

        def _unique_name_with_suffix(base_sheet_name: str, initial_suffix: int | None = None) -> str:
            """Generate a unique sheet name with a numeric suffix."""
            unique_sheet_name = base_sheet_name
            suffix_n = len([s for s in writer_sheet_names_uc if s == unique_sheet_name.upper()]) + 1

            if initial_suffix is not None:
                suffix = str(initial_suffix)
                unique_sheet_name = base_sheet_name[: WORKSHEET_NAME_LEN - len(suffix)] + suffix

            while unique_sheet_name.upper() in writer_sheet_names_uc:
                suffix = str(suffix_n)
                unique_sheet_name = base_sheet_name[: WORKSHEET_NAME_LEN - len(suffix)] + suffix
                suffix_n += 1

            return unique_sheet_name

        safe_sheet_name = truncate_w_ellipsis(requested_name, WORKSHEET_NAME_LEN)

        for char in WORKSHEET_NAME_ILLEGAL_CHARS:
            safe_sheet_name = safe_sheet_name.replace(char, "_")

        base_sheet_name = safe_sheet_name

        if not base_sheet_name:
            # Unexpected but no reason to fail
            base_sheet_name = "Sheet"
            safe_sheet_name = _unique_name_with_suffix(base_sheet_name, 1)
            assert safe_sheet_name.upper() not in writer_sheet_names_uc

        if safe_sheet_name.upper() in writer_sheet_names_uc:
            use_existing_sheet = self.worksheet_config.get("use_existing", False)

            if use_existing_sheet:
                if safe_sheet_name not in self.workbook.can_use_existing_sheet:
                    logger.warning(f"Worksheet name '{safe_sheet_name}' is not in template so cannot be reused; renaming")
                    use_existing_sheet = False  # rename sheet
                elif not self.workbook.can_use_existing_sheet[safe_sheet_name]:
                    logger.warning(f"Worksheet name '{safe_sheet_name}' already reused once; renaming")
                    use_existing_sheet = False  # rename sheet
                else:
                    self.workbook.can_use_existing_sheet[safe_sheet_name] = False  # only use existing sheet once

            if not use_existing_sheet:
                safe_sheet_name = _unique_name_with_suffix(base_sheet_name)
                assert safe_sheet_name.upper() not in writer_sheet_names_uc

        assert len(safe_sheet_name) <= WORKSHEET_NAME_LEN

        if safe_sheet_name != requested_name:
            logger.debug(f"Worksheet name '{requested_name}' renamed to '{safe_sheet_name}'")

        return safe_sheet_name

    def _get_header_letter_idx(
        self,
        header: DataFrameKeyT | None = None,
        letter: str | None = None,
        idx: int | None = None,
        cell: Cell | MergedCell | None = None,
        raise_missing_header: bool = True,
    ) -> int | None:
        """Return a column index given an index, header or letter.

        Fails until `to_excel` is complete.
        Requires worksheet to have at least one row.

        Args:
            header (DataFrameIndexHashableT): Column header value; None if `letter` or `idx` is used.
            letter (str): Column letter (A, B, ...); None if `header` or `idx` is used.
            idx (str): Column index; (1=A, 2=B, ...); None if `header` or `letter` is used.
            cell (Cell | MergedCell): Cell reference; None if `header`, `letter` or `idx` is used.
            raise_missing_header (bool): True to raise `ValueError` if header value cannot be found.

        Returns:
            int | None
        """
        assert self.pyxl_worksheet is not None

        if idx is not None:
            return idx
        if cell is not None:
            return cell.column
        if letter is not None:
            return column_index_from_string(letter)
        if header is not None:
            letter = self.header_to_letter_map.get(header)
            if letter is not None:
                return column_index_from_string(letter)
            if raise_missing_header:
                raise ValueError(f"Worksheet header '{header}' not found.")

        return None

    @classmethod
    def _get_row_values(
        cls,
        row_cells: Sequence[Cell | MergedCell],
        expand_merged: bool = True,
    ) -> Sequence[WorksheetCellValueT | None]:
        """Return a list of values for all cells in a row.

        Non-merged cells with `None` value are replaced with an empty string.

        Args:
            row_cells (Sequence[Cell | MergedCell]): Row cells.
            expand_merged (bool): True to set values of merged cells; otherwise merged cells will have value `None`.

        Returns:
            list[str]
        """
        row_values: list[WorksheetCellValueT | None] = []

        for cell in row_cells:
            if isinstance(cell, MergedCell):
                # The first value in a set of merged cells is the value of the merged cell.
                row_values.append(row_values[-1] if expand_merged else None)
            else:
                assert isinstance(cell, Cell)
                row_values.append(cell.value if cell.value is not None else "")

        assert len(row_values) == len(row_cells)

        return row_values
